import openpyxl

def readUsernameData():
    book = openpyxl.load_workbook("C:\\Users\\user\\OneDrive\\Desktop\\py_sel.xlsx")
    sheet = book.active
    if "LoginPage" in book.sheetnames:
        sheet = book.worksheets[0]
        username = sheet.cell(row=3, column=1).value
        # print(username)
    return username
print(readUsernameData())