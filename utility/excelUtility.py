import openpyxl

class ExcelUtility:
    def read_user_data(self,rownum,colnum):
        book = openpyxl.load_workbook("C:\\Users\\user\\Downloads\\TestData.xlsx")
        sheet = book.active
        if "LoginPage" in book.sheetnames:
            sheet = book.worksheets[0]
            username = sheet.cell(row=rownum, column=colnum).value
            # print(username)
        return username
